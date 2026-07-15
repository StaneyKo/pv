"""Excel 材料解析器。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from models import SourceChunk


def parse_excel(file_path: Path) -> list[SourceChunk]:
    """按工作表和有效行提取 Excel 内容。"""
    workbook = load_workbook(str(file_path), read_only=True, data_only=True)
    chunks: list[SourceChunk] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[str] = []
            start_row: int | None = None
            end_row: int | None = None
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    start_row = start_row or row_number
                    end_row = row_number
                    rows.append(" | ".join(values))
            if rows:
                location = f"工作表“{worksheet.title}” 行 {start_row}-{end_row}"
                chunks.append(SourceChunk("\n".join(rows), file_path.name, location, "Excel"))
    finally:
        workbook.close()
    return chunks

